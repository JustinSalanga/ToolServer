import re
from urllib.parse import urlparse, parse_qs, unquote, urlunparse, urlencode, parse_qsl

def extract_target_url(url):
    parsed = urlparse(url)
    
    # Handle known redirect patterns
    redirect_domains = {
        'www.indeed.com': 'jk',
        'www.wiraa.com': 'source'
    }
    
    if parsed.netloc in redirect_domains:
        query = parse_qs(parsed.query)
        target_param = redirect_domains[parsed.netloc]

        # For Indeed, you might not get the full job URL, just job key
        if redirect_domains[parsed.netloc] in query:
            return f"https://{parsed.netloc}/viewjob?jk={query[redirect_domains[parsed.netloc]][0]}"

        if target_param in query:
            # Decode the redirect target
            return unquote(query[target_param][0])
    
    # If not a known redirector, return the original
    return urlunparse(parsed._replace(query=normalize_query(parsed.query)))

def normalize_query(query_string):
    query_pairs = sorted(parse_qsl(query_string))
    return urlencode(query_pairs)

def normalize_final_url(url):
    parsed = urlparse(url)
    
    netloc = parsed.netloc.lower().replace("www.", "")
    path = parsed.path.rstrip("/")
    query = normalize_query(parsed.query)
    
    return urlunparse((parsed.scheme, netloc, path, '', query, ''))

def are_job_urls_same(url1, url2):
    u1 = normalize_final_url(extract_target_url(url1))
    u2 = normalize_final_url(extract_target_url(url2))
    return u1 == u2

import pandas as pd
# Set options to prevent wrapping
pd.set_option('display.width', None)         # Don't limit line width
pd.set_option('display.max_columns', 5)   # Show all columns
pd.set_option('display.max_colwidth', None)  # Don't truncate column content
df = pd.read_excel('bid.xlsx', sheet_name='Main')
# print(df.iloc[i])
# check_list = 'COMPANY NAME'
check_list = 'LINK'
link_list = {}

resume_cnt_dict = {}

def clearLink(link):
	link = str(link)
	if '?' in link and (not 'indeed' in link) and (not 'builtin' in link) and (not 'wellfound' in link) and (not 'wiraa' in link):
		link = link.split('?')[0]
	if link[-1] == '/':
		link = link[:-1]
	if '/apply' == link[-6:]:
		link = link[:-6]
	if '/application' == link[-12:]:
		link = link[:-12]

	if 'www.indeed.com' in link:
		link = extract_target_url(link)
		# print(link)

	return link

# print(clearLink('https://jobs.ashbyhq.com/worldly/6b2b92c8-5bc6-432a-91a1-5e5fb6fc1ad4/application/apply?'))\

all_links = []
for i in range(len(df)):
	all_links.append((str(df.iloc[i]['NO']),str(df.iloc[i]['LINK'])))

# for i in range(len(all_links)):
# 	for k in range(len(all_links)):
# 		if i != k:
# 			a, b = all_links[i][1], all_links[k][1]
# 			if are_job_urls_same(a, b):
# 				print(all_links[i][0], all_links[k][0], a, b)
# 				print('------------')
# print("=============")
duplicated_ids = []
for i in range(len(df)):
	row = df.iloc[i]
	resume = row['RESUME']
	if resume:
		resume = str(resume)
		resume = resume.replace(' ', '')
		if not resume in resume_cnt_dict:
			resume_cnt_dict[resume] = 1
		else:
			resume_cnt_dict[resume] += 1
	dIndex = clearLink(row[check_list])
	if not dIndex in link_list:
		link_list[dIndex] = [row['NO']]
	else:
		link_list[dIndex].append(row['NO'])
		print(row['NO'], row['COMPANY NAME'], '\t\t', row['JOB TITLE'], )
		duplicated_ids.append(str(row['NO']))


resume_list = []

for k in resume_cnt_dict:
	resume_list.append((k, resume_cnt_dict[k]))



resume_list = sorted(resume_list, key = lambda x: x[1], reverse = True)

# print('len: ', len(resume_list))

# for i in resume_list:
# 	print(i[0], '\t\t\t', i[1])

print(duplicated_ids)


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the worksheet
wb = load_workbook('bid.xlsx')
ws = wb['Main']  # Change to your sheet name

# Define fill color
highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
ghighlight = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Yellow

# Loop through specific cells and apply style
for row in ws.iter_rows():  # Change range as needed
	if str(row[0].value) in duplicated_ids:
	    for cell in row:
	        cell.fill = highlight
	else:
		for cell in row:
			cell.fill = ghighlight

# Save the file
wb.save('bid.xlsx')
